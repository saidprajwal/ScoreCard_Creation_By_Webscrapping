import requests, webbrowser
from bs4 import BeautifulSoup
import openpyxl
import json


wb=openpyxl.load_workbook("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\StartUp.xlsx")
sh1=wb['Sheet1']
startUps=[]
for i in range (2,sh1.max_row+1):
    if (sh1.cell(row=i, column=2).value == None):
        break
    startUps.append(sh1.cell(row=i,column=2).value)

myjsonfile=open("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Passion_Process.json","r")
jsondata=myjsonfile.read()

obj=json.loads(jsondata)

li=obj['Process_areas']
process_areas=[]
for i in range(len(li)):
    process_areas.append(li[i].get("Process"))
#print(process_areas)

i=1
for list in startUps:
 rows = 1
 col = 1
 i = i + 1
 s = 'Sheet' + str(i)
 rows=rows+1
 sh2 = wb[s]
 c1 = sh2.cell(row=rows, column=col)
 c1.value = list
 rows=rows+1
 wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\StartUp.xlsx")
 for value in process_areas:
    col=1
    c1 = sh2.cell(row=rows, column=col)
    c1.value = value
    wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\StartUp.xlsx")
    col = col + 1
    print(list+" "+ value + " " + "Management")
    google_search = requests.get("https://www.google.co.in/search?q=" + list+" "+ value )
    htmlContent = google_search.content
    soup = BeautifulSoup(htmlContent, 'html.parser')
    anchors = soup.find_all('a')
    for link in anchors:
        a = link.get('href')
        if (link.get('href')[0:4] == '/url'):
            print(link.get('href'))
            c=sh2.cell(row=rows,column=col)
            c.value=a
            col=col+1
    rows=rows+1


print("-----------------------------------------------------------------------------------\n")

wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\StartUp.xlsx")

