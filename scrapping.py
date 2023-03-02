import requests, webbrowser
from bs4 import BeautifulSoup
import openpyxl
import json


wb=openpyxl.load_workbook("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
sh1=wb['Sheet1']
startUps=[]
for i in range (2,sh1.max_row+1):
    if (sh1.cell(row=i, column=2).value == None):
        break
    startUps.append(sh1.cell(row=i,column=2).value)

myjsonfile=open("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Passion_Process.json","r")
jsondata=myjsonfile.read()

obj=json.loads(jsondata)

li=obj['Process_areas']
process_areas=[]
weight=[]
maxscore=[]
for i in range(len(li)):
    process_areas.append(li[i].get("Process"))
    weight.append(li[i].get("Weight"))
    maxscore.append(li[i].get("Maxscore"))



print(process_areas)
print(weight)
print(maxscore)
count=0
i=1
weight_count=0
for list in startUps:
 weight_count = 0
 rows = 1
 col = 1
 i = i + 1
 s = 'Sheet' + str(i)
 rows=rows+1
 sh2 = wb[s]
 c1 = sh2.cell(row=rows, column=col)
 c1.value = list
 wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
 rows=rows+1
 col = 1
 c1 = sh2.cell(row=rows, column=col)
 c1.value = "Process Area"
 wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
 col = 2
 c1 = sh2.cell(row=rows, column=col)
 c1.value = "Links"
 wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
 col = 3
 c1 = sh2.cell(row=rows, column=col)
 c1.value = "Link Count"
 wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
 col = 4
 c1 = sh2.cell(row=rows, column=col)
 c1.value = "Weight"
 wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
 col = 5
 c1 = sh2.cell(row=rows, column=col)
 c1.value = "Maxscore"
 wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
 col = 6
 c1 = sh2.cell(row=rows, column=col)
 c1.value = "Actual Score"
 wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
 col = 7
 c1 = sh2.cell(row=rows, column=col)
 c1.value = "Remark"
 wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
 for value in process_areas:
    count=0
    col=1
    rows = rows + 1
    row_count =rows
    c1 = sh2.cell(row=rows, column=col)
    c1.value = value
    wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
    col=col+1
    print(list+" "+ value + " " + "Management")
    google_search = requests.get("https://www.google.co.in/search?q=" + list+" "+ value )
    htmlContent = google_search.content
    soup = BeautifulSoup(htmlContent, 'html.parser')
    anchors = soup.find_all('a')
    for link in anchors:
        a = link.get('href')
        if (link.get('href')[0:4] == '/url'):
            print(link.get('href'))
            count=count+1
            c=sh2.cell(row=rows,column=col)
            c.value=a
            rows=rows+1
    c1 = sh2.cell(row=row_count, column=3)
    c1.value = count
    wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
    j=row_count
    for k in range(1,count):
        j=j+1
        c1 = sh2.cell(row=j, column=3)
        c1.value = " "
        wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
    j=row_count
    for k in range(0,count):
        c1 = sh2.cell(row=j, column=4)
        c1.value = weight[weight_count]
        wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
        j=j+1
    j = row_count
    for k in range(0, count):
        c1 = sh2.cell(row=j, column=5)
        c1.value = maxscore[weight_count]
        wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
        j = j + 1
    weight_count = weight_count + 1




print("-----------------------------------------------------------------------------------\n")

wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")

