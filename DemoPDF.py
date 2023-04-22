
import PyPDF2
import openpyxl
file="C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx"
wb=openpyxl.load_workbook(file)
sh2=wb['Sheet2']

reader=PyPDF2.PdfReader("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\CAV.pdf")
pages=len(reader.pages)
str=""

break_point=1
#store scores from excel in list
scores=[]
for i in range (2,sh2.max_row+1):
    if(sh2.cell(row=i,column=1).value==None):
         break_point=i
         break
    scores.append(sh2.cell(row=i,column=1).value)
#print(scores)

#check if score is present in pdf and store value
total_score=0
rows=2
flag=0
for item in scores:
    c = sh2.cell(row=rows, column=2)
    for i in range(0,pages):
        str=reader.pages[i].extract_text()
        if item.lower() in str.lower():
            flag=1
            c.value = 1;
            total_score+=1;
            wb.save(file)
            break
    rows=rows+1
    if flag==0:
        c.value = 0
        wb.save(file)

#store total score obtained in break_point row
c=sh2.cell(row=break_point,column=2)
c.value="{}/{}".format(total_score,len(scores))
wb.save(file)



