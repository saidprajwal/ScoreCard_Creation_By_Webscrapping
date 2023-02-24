import openpyxl

wb=openpyxl.load_workbook("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
print(type(wb))

#Name of sheets in workbook
print(wb.sheetnames)

#current active sheet
print(wb.active.title)

#access data cell by cell
sh1=wb['Sheet1']
data=sh1['B2'].value    #data=wb['Sheet1']['B2'].values
print(data)

#access data by row cell and column cell only one
print(sh1.cell(2,2).value)

print(sh1.cell(row=1,column=2).value)

#print column
# for i in range (2,sh1.max_row+1):
#     print(sh1.cell(row=i,column=2).value)

c = sh1.cell(row=15, column=3)
c.value =213546



wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")
c = sh1.cell(row=16, column=3)
c.value =213546
wb.save("C:\\Users\\PRAJWAL\\PycharmProjects\\WebScapping\\excel\\Audit.xlsx")